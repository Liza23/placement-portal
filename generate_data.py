# Author: Aditya Badola
# Aim: Generate dummy Master Data for Skynet Simulation | Generate data for Database Management Project
# Written On: 13-11-2021 11:00 PM IST
# Updated On: 20-04-2022 8:00 PM IST

from __future__ import print_function
import mailbox
from operator import inv
from eth_typing import ContractName
import numpy as np
from openpyxl import load_workbook
from shutil import copy2, make_archive
import atexit, os
from os import path
from json import dumps, loads



import random, string

# ENTITY 1 : COORDINATOR ==============================

# define a class for coordinators
class Coordie:
    def __init__(self, name, mail, cntct):
        self.name = name
        self.mail = mail
        self.cntct = cntct

coordies = [Coordie("Aditya Badola", "190050006@iitb.ac.in", "8529295025"),
            Coordie("Advait Kumar", "18D070003@iitb.ac.in", "9773817417"),
            Coordie("Anurag Kr Gupta", "203186001@iitb.ac.in", "9136824962"),
            Coordie("Baksheesh Sachar", "21f468004@iitb.ac.in", "9667640017"),
            Coordie("Jasleen Kaur", "215280046@iitb.ac.in", "9855591548"),
            Coordie("Nikhil Parpay", "213100079@iitb.ac.in", "8275233411"),
            Coordie("Pradeep Chudsama", "21370111@iitb.ac.in", "9354816573"),
            Coordie("Ravindra Girase", "213020004@iitb.ac.in", "8600154174"),
            Coordie("Sanjna Mohan", "20305r006@iitb.ac.in", "9847051590"),
            Coordie("Suyash Kadam", "190020118@iitb.ac.in", "9307135660"),
            Coordie("Kunal Jain", "kunaljain@iitb.ac.in", "7734006476", ),
            Coordie("Tushar Dhingra", "190110099@iitb.ac.in", "7999004202"),
            Coordie("Pushpendra Singh Jadoun", "190040082@iitb.ac.in", "8079046720"),
            Coordie("Tarunsai Goddu", "190070024@iitb.ac.in", "6304713048"),
            Coordie("Utkarsh Sahu", "180100120@iitb.ac.in", "6263905277"),
            Coordie("Ayush Aditya", "21f468009@iitb.ac.in", "7093195305"),
            Coordie("Govind Rajhans Jadhav", "20307r004@iitb.ac.in", "8237566400"),
            Coordie("Pasunuri Prathiba", "203070024@iitb.ac.in", "6301832970"),
            Coordie("Rajib Chatterjee", "20307r005@iitb.ac.in", "7003035535"),
            Coordie("Mohammed Riyaz", "213110053@iitb.ac.in", "9480216652")]
coordies.sort(key=lambda x: x.name)
pswds = [''.join(random.choice(string.ascii_uppercase + string.ascii_lowercase + string.digits) for _ in range(16)) for i in range(len(coordies))]

# write the data of coordinators to a .csv file
f = open("coordinators.csv", 'w')
f.write("coordinator_id, coordinator_password, coordinator_name, coordinator_email, coordinator_contact\n")
for i in range(len(coordies)):
    f.write(f"{i+1}, {pswds[i]}, {coordies[i].name}, {coordies[i].mail}, {coordies[i].cntct}\n")
f.close()


# ENTITY 2 : PROGRAM ==============================

# define a class for programs
class Prog:
    def __init__(self, id, name, dur, lvl):
        self.id = id
        self.name = name
        self.dur = dur
        self.lvl = lvl
 
progs = [Prog(1, "B.Tech.", 4, "UG"),
         Prog(2, "Dual Degree", 5, "UG"),
         Prog(3, "B.S.", 4, "UG"),
         Prog(4, "B.Des.", 4, "UG"),
         Prog(5, "M.Sc.", 2, "PG"),
         Prog(6, "M.Tech. (TA)", 2, "PG"),
         Prog(7, "M.Tech. (RA)", 3, "PG"),
         Prog(8, "M.Des.", 2, "PG"),
         Prog(9, "M.Phil.", 2, "PG"),
         Prog(10, "MPP", 2, "PG"),
         Prog(11, "M.S. by Research", 2, "PG"),
         Prog(12, "Ph.D.", 6, "PhD")]

# write the data of programs to a .csv file
f = open("programs.csv", 'w')
f.write("program_id, program_name, program_duration, program_level\n")
for prog in progs:
    f.write(f"{prog.id}, {prog.name}, {prog.dur}, {prog.lvl}\n")
f.close()


# ENTITY 3 : DEPARTMENT ==============================

# define a class for departments
class Dept:
    def __init__(self, name, code):
        self.name = name
        self.code = code

depts = [Dept("Aerospace Engineering", "AE"), 
         Dept("Biosciences and Bioengineering", "BM"),
         Dept("Chemical Engineering", "CH1"), 
         Dept("Chemistry", "CH2"), 
         Dept("Civil Engineering", "CE"), 
         Dept("Computer Science & Engineering", "CS"), 
         Dept("Earth Sciences", "ES"), 
         Dept("Electrical Engineering", "EE"), 
         Dept("Energy Science and Engineering", "EN"), 
         Dept("Environmental Science and Engineering", "EV"), 
         Dept("Humanities & Social Sciences", "HSS"), 
         Dept("Mathematics", "MA"), 
         Dept("Mechanical Engineering", "ME"), 
         Dept("Metallurgical Engineering & Materials Science", "MM"), 
         Dept("Physics", "PH"), 
         Dept("IDC School of Design", "IDC"), 
         Dept("Shailesh J. Mehta School of Management", "SJSOM"), 
         Dept("Geoinformatics & Natural Resources Engineering", "GNR"), 
         Dept("Industrial Engineering and Operations Research", "IO"), 
         Dept("Materials, Manufacturing and Modeling", "MMM"), 
         Dept("Systems & Control Engineering", "SC"), 
         Dept("Centre for Technology Alternatives for Rural Areas", "CTARA"), 
         Dept("Centre for Machine Intelligence and Data Science", "C-MInDS"), 
         Dept("Centre of Studies in Resources Engineering", "CSRE")]
depts.sort(key=lambda x: x.name)

# write the data of departments to a .csv file
f = open("departments.csv", 'w')
f.write("department_id, department_name\n")
for dept in depts:
    f.write(f"{dept.code}, {dept.name}\n")
f.close()


# ENTITY 4 : PROFILE ==============================

# define a class for profiles
class Profile:
    def __init__(self, name, descrip):
        self.name = name
        self.descrip = descrip

profs = [Profile("Consultant", "Consulting covers an incredibly broad range of topics, organizations, clients, and industries. A consultant can work independently or as part of a consulting firm, specializing in any given field of expertise. Essentially, consultants are hired to share their expertise and knowledge to help businesses attain goals and solve problems. Sometimes, companies bring on consultants to perform day-to-day work and augment or supplement staff—and save the fixed overhead costs associated with a full-time employee. Other times, they're brought on board for a specialized purpose: to troubleshoot, tackle a specific challenge, optimize something specific, spin off a successful business unit, revive a failing business, etc."),
         Profile("Business Analyst", "Business analysts play a key role in the success of so many organizations. Acting as communicators, facilitators, and mediators, these flexible professionals seek out the best ways to improve processes and increase effectiveness through technology, strategy, analytic solutions, and more. With organizational and project goals lighting the way, business analysts are meticulous in their documentation and evaluation of potential solutions—always working to bridge the gap among departments with improved technical efficiency and productivity."),
         Profile("Operations", "Operations managers oversee the organizational activities of businesses, government agencies, non-profit groups, and other organizations. These professionals are talented managers and leaders. They might support operational leadership in a variety of departments — from finance and IT to human resources and accounts payable. At both large and small organizations, operations managers supervise, hire, and train employees, manage quality assurance programs, strategize process improvements, and more. Operations managers are ultimately responsible for maintaining and increasing the efficiency of a business, agency, or organization."),
         Profile("Associate Product Manager", "Within many industries, product managers are responsible for guaranteeing the success of a specific product or product line. These extreme doers are product experts with a powerful capability to make strategic decisions based on market and competitor analyses. Prior to production, they create the roadmap to guide a product from conception through design and into wide release. Product managers bridge the gap among the different departments involved in successfully bringing your product(s) to market, including R&D, engineering, manufacturing, marketing, sales, and customer support. Their ultimate goal is the creation and launch of products that meet consumers' needs — growing market share and success."),
         Profile("Finance", "Financial analysts are fundamental contributors to the fiscal health and success of many types of organizations. From banks and pension and mutual funds, to securities firms and insurance companies, these highly adept people evaluate economic data and trends, determine financial status and value, and help guide investment decisions. Financial analysts — sometimes referred to as securities or investment analysts — often specialize in specific industries, products, or geographic regions. On both the buy-side and sell-side of the financial landscape, they may work as portfolio or fund managers, ratings analysts, risk analysts, and more."),
         Profile("Data Science/Analytics", "A data scientist knows how to extract meaning from and interpret data. This unique skill set requires the aid of statistical methods and machinery, but largely relies on analytical brain power. Because raw data can rarely be utilized reliably, businesses in a variety of industries look to these technical experts to collect, clean, and validate their data. This meticulous process requires persistence and software engineering skills—expertise that's integral to understanding data bias and to debug output from code. In simpler terms, data scientists find patterns and use the knowledge to build and improve."),
         Profile("Quantitative Trading", "Quantitative trading (also called quant trading) involves the use of computer algorithms and programs—based on simple or complex mathematical models—to identify and capitalize on available trading opportunities. Quant trading also involves research work on historical data with an aim to identify profit opportunities."),
         Profile("IT Software", "In our digital world, organizations have to stay connected at all costs. The role of information technology, or IT, is integral to implementing and maintaining the infrastructure and solutions that will continue to move the business forward. The IT department is there to assist as computer issues arise, software needs updating, or networks require fixing. In general, the IT department is responsible for implementing infrastructure automation, governing the use of network and operating systems, and optimizing functionality. At one point or another, he or she will typically save the day by recovering a crucial document—or preventing a systemwide cybersecurity breach, or keeping such a breach from spreading."),
         Profile("Core", "Department-wise core profile"),
         Profile("FMCG", "FMCG covers a range of sectors and job titles within retail. You could work as a Sales Assistant on the shop floor, boss people around in Management or work behind the scenes as a Buyer or Merchandiser. Each of these has very difficult responsibilities, but can all sell FMCG. Although daily duties are different for every role you're likely to start out as a Sales Assistant on the shop floor. Spending a lot of time on your feet, the job involves: Serving customers, Displaying products, Overseeing deliveries, Handling payments, Helping with special promotions"),
         Profile("Design", "Depending on the medium, a designer might use two-dimensional, three-dimensional, and/or graphic art to create visual concepts and products. Layout design, formatting, organization, and prioritization of materials, topics, images, and text are frequently part of their work; often, the resulting concepts or products aim to communicate a central message or theme. Designers work across a wide range of industries, disciplines, and roles, from in-house to agency or freelance design. They also work with an extensive variety of media, tools, techniques, and technologies.")]
idToProfile = dict()

# write the data of profiles to a .csv file
f = open("profiles.csv", 'w')
f.write("profile_id\tprofile_name\tprofile_description\n")
for i in range(len(profs)):
    idToProfile[i+1] = profs[i].name
    f.write(f"{i+1}\t{profs[i].name}\t{profs[i].descrip}\n")
f.close()
profileToId = {v: k for k, v in idToProfile.items()}


# ENTITY 5 : COMPANY ==============================

# countries of origin
C = ["India", "United States", "Japan", "Hong Kong", "South Korea"]

firmNames = []
f1 = open("firmNameDataset.txt", 'r')
for line in f1:
    firmNames.append(line.strip())
f1.close()
idToFirm = dict()
numCompanies = len(firmNames)

# write the data of companies to a .csv file
f2 = open("companies.csv", 'w')
f2.write("company_id, company_name, company_origin, company_coordinator\n")
i = 0
for i in range(numCompanies):
    idToFirm[i+1] = firmNames[i]
    r = np.random.random()
    if r < 0.85:
        f2.write(f"{i+1}, {firmNames[i]}, {C[0]}, {i%20 + 1}\n")
    elif r < 0.9:
        f2.write(f"{i+1}, {firmNames[i]}, {C[1]}, {i%20 + 1}\n")
    elif r < 0.95:
        f2.write(f"{i+1}, {firmNames[i]}, {C[2]}, {i%20 + 1}\n")
    elif r < 0.975:
        f2.write(f"{i+1}, {firmNames[i]}, {C[3]}, {i%20 + 1}\n")
    else:
        f2.write(f"{i+1}, {firmNames[i]}, {C[4]}, {i%20 + 1}\n")
    i += 1
f2.close()
firmToId = {v: k for k, v in idToFirm.items()}


# UTILITY FUNCTIONS : randomised contacts and names ==============================

def randPhNo():
    ph_no = ""
    ph_no += str(np.random.randint(7,10))
    for i in range(1,10):
        ph_no += str(np.random.randint(0,10))
    return ph_no

firstNames = []
f1 = open("firstNameDataset.txt", 'r')
for line in f1:
    firstNames.append(line.strip().split()[0])
f1.close()

lastNames = []
f2 = open("lastNameDataset.txt", 'r')
for line in f2:
    lastNames.append(line.strip().split()[0])
f2.close()


# ENTITY 6 : RECRUITER ==============================

numRecrtrs = int(numCompanies*1.05)
recrtrs = []
pswds = [''.join(random.choice(string.ascii_uppercase + string.ascii_lowercase + string.digits) for _ in range(16)) for _ in range(numRecrtrs)]
rcrtrFirstName = np.random.choice(firstNames, numRecrtrs)
rcrtrLastName = np.random.choice(lastNames, numRecrtrs)
cntcts = [randPhNo() for _ in range(numRecrtrs)]

I = [i for i in range(numRecrtrs)]
random.shuffle(I)

f3 = open("recruiters.csv", 'w')
f3.write("recruiter_id, recruiter_password, recruiter_name, recruiter_contact, recruiter_email, recruiter_company\n")
for i in range(numRecrtrs):
    f3.write(f"{i+1}, {pswds[i]}, {rcrtrFirstName[i]} {rcrtrLastName[i]}, {cntcts[i]}, {rcrtrFirstName[i].lower()}.{rcrtrLastName[i].lower()}{np.random.randint(1,99)}@{firmNames[I[i]%numCompanies].lower().split()[0]}.com, {I[i]%numCompanies+1}\n")
f3.close()


# ENTITY 7 : JAF ==============================

# define a class for jaf
class JAF:
    def __init__(self, company_id, title, profile_id):
        self.cid = company_id
        self.title = title
        self.pid = profile_id

jafs = []
f1 = open("jobs.tsv", 'r')
r = 0
for line in f1:
    if r != 0:
        jaf = line.strip().split('\t')
        jafs.append(JAF(firmToId[jaf[0]], jaf[1], profileToId[jaf[2]]))
    r += 1
f1.close()

# def read_counter():
#     return loads(open("counter.json", "r").read()) + 1 if path.exists("counter.json") else 0

# def write_counter():
#     with open("counter.json", "w") as f:
#         f.write(dumps(counter))

# counter = read_counter()
# atexit.register(write_counter)


# def randGender():
#     p = np.random.rand()
#     if p > 0.5:
#         return "M"
#     else:
#         return "F"

# F = ['1 StudentTemplate.xlsx', '2 APC Template.xlsx', '3 ComPOCTemplate.xlsx', '4 IRFTemplate.xlsx', '5 RoundTemplate.xlsx', '6 InterviewShortlistTemplate.xlsx', '9.CommonVolunteerMappingTemplate.xlsx']

# for i in range(len(F)):
#     copy2('/home/badola/APC/IIT Bombay Data Templates/' + F[i], F[i])

# N = 70        # total students
# C = 10          # total firms
# C_1 = 10
# C_2 = 0
# C_3 = C - C_1 - C_2
# numAPC = 20     # total APCs
# numPOC = 36     # total ComPOCs

# f3 = open("firmNameDataset.txt", 'r')

# C1 = []
# C2 = []
# C3 = []


# wb = load_workbook(filename='IPT Contacts 2021-22.xlsx')
# ws = wb['ICs']
# for i in range(2, 2+numPOC):
#     C1.append(ws.cell(row=i, column=1).value)
#     C2.append(ws.cell(row=i, column=7).value)
#     C3.append(ws.cell(row=i, column=4).value)
# wb.save('IPT Contacts 2021-22.xlsx')


# br = "--------------------------------------------------------------------"
# print(br)
# print("Number of total students: ", N)
# print("Number of firms: ", C)
# print(br)

# # Student Details
# firstName = np.random.choice(D1, N)
# lastName = np.random.choice(D2, N)
# email = []
# regNum = []
# contactNum = []
# gender = []
# for i in range(N):
#     regNum.append("iitb_" + str(i))
#     email.append(firstName[i].lower() + "." + lastName[i].lower() + "@iitb.ac.in")
#     contactNum.append(randPhNo())
#     gender.append(randGender())

# print("\nStarting to write in .xlsx files...\n\n")


# # 1 StudentTemplate

# wb = load_workbook(filename = F[0])
# ws = wb['Sheet1']
# for i in range(2, 2+N):
#     ws.cell(row=i, column=1).value = firstName[i-2] + " " + lastName[i-2]
#     ws.cell(row=i, column=2).value = email[i-2]
#     ws.cell(row=i, column=3).value = regNum[i-2]
#     ws.cell(row=i, column=4).value = contactNum[i-2]
#     ws.cell(row=i, column=5).value = gender[i-2]
# wb.save(F[0])


# # 2 APC Template

# wb = load_workbook(filename = F[1])
# ws = wb['Sheet1']
# ws.cell(row=i, column=1).value = C1[0]
# ws.cell(row=i, column=2).value = C2[0]
# ws.cell(row=i, column=3).value = C3[0]
# wb.save(F[1])


# # 3 ComPOCTemplate

# wb = load_workbook(filename = F[2])
# ws = wb['Sheet1']
# for i in range(2, 2+numPOC):
#     ws.cell(row=i, column=1).value = C1[i-2]
#     ws.cell(row=i, column=2).value = C2[i-2]
#     ws.cell(row=i, column=3).value = C3[i-2]
# wb.save(F[2])


# random.shuffle(C2)

# # 4 InterviewTemplate

# firmName = np.random.choice(D3, C, replace=False)
# roles1 = ["Consult", "BA", "Finance", "APM", "FMCG"]
# roles2 = ["DS", "Analytics", "Quant", "IT"]
# roles3 = ["Operations", "Core"]
# firmEmail = []
# jobRole = []
# pocEmail = []
# numPanels = []
# numRounds = []
# J = 0
# for i in range(C_1):
#     firmEmail.append(firmName[i].strip().split()[0].lower() + "@gmail.com")
#     jobRole.append(roles1[np.random.randint(0,5)])
#     pocEmail.append(C2[i+J])
#     # pocEmail.append(C2[np.random.randint(0,numAPC)])
#     # numPanels.append(3)
#     numPanels.append(np.random.randint(3,4))
#     numRounds.append(np.random.randint(2,4))
# J += C_1
# for i in range(C_2):
#     firmEmail.append(firmName[i].strip().split()[0].lower() + "@gmail.com")
#     jobRole.append(roles2[np.random.randint(0,4)])
#     pocEmail.append(C2[i+J])
#     # pocEmail.append(C2[np.random.randint(0,numAPC)])
#     # numPanels.append(3)
#     numPanels.append(np.random.randint(3,4))
#     numRounds.append(np.random.randint(2,4))
# J += C_2
# for i in range(C_3):
#     firmEmail.append(firmName[i].strip().split()[0].lower() + "@gmail.com")
#     jobRole.append(roles3[np.random.randint(0,2)])
#     pocEmail.append(C2[i+J])
#     # pocEmail.append(C2[np.random.randint(0,numAPC)])
#     # numPanels.append(3)
#     numPanels.append(np.random.randint(3,4))
#     numRounds.append(np.random.randint(2,4))

# wb = load_workbook(filename = F[3])
# ws = wb['Sheet1']
# for i in range(2, 2+C):
#     ws.cell(row=i, column=1).value = firmName[i-2]
#     ws.cell(row=i, column=2).value = firmEmail[i-2]
#     ws.cell(row=i, column=3).value = jobRole[i-2]
#     ws.cell(row=i, column=4).value = pocEmail[(i-2)]
#     ws.cell(row=i, column=5).value = X2[(i-2)%20]
#     ws.cell(row=i, column=6).value = numPanels[i-2]
#     ws.cell(row=i, column=7).value = "2021-11-30"
#     # ws.cell(row=i, column=8).value = "01:30:00"
#     k = i-2
#     if 0 <= k and k < C_1:
#         ws.cell(row=i, column=9).value = np.random.choice([10])
#         ws.cell(row=i, column=8).value = "16:45:00"
#     elif k < (C_1 + C_2):
#         ws.cell(row=i, column=9).value = np.random.choice([10])
#         ws.cell(row=i, column=8).value = "01:50:00"
#     else:
#         ws.cell(row=i, column=9).value = np.random.choice([10])
#         ws.cell(row=i, column=8).value = "02:10:00"
#     tmp = "PI Round 1"
#     for j in range(numRounds[i-2]):
#         tmp += ","
#         tmp += "PI Round " + str(j+2)
#     ws.cell(row=i, column=10).value = tmp
#     # ws.cell(row=i, column=13).value = "PI Round 1"
#     # ws.cell(row=i, column=14).value = "0"
# wb.save(F[3])


# # 5 RoundTemplate

# wb = load_workbook(filename = F[4])
# ws = wb['Sheet1']
# curr_row = 2
# for i in range(2, 2+C):
#     # p = 3
#     p = np.random.randint(2,5)
#     for j in range(numRounds[i-2]):
#         ws.cell(row=curr_row, column=1).value = firmName[i-2] + "(" + jobRole[i-2] + ")"
#         ws.cell(row=curr_row, column=2).value = str(j+1)
#         ws.cell(row=curr_row, column=3).value = "PI Round " + str(j+1)
#         ws.cell(row=curr_row, column=4).value = "0"
#         curr_row += 1
# wb.save(F[4])


# # 6 InterviewShortlistTemplate

# J = 0

# numShortlists = []
# for i in range(C):
#     # numShortlists.append(12)
#     if 0 <= i and i < C_1:
#         numShortlists.append(np.random.randint(7,20))
#     elif i < (C_1 + C_2):
#         numShortlists.append(np.random.randint(6,10))
#     else:
#         numShortlists.append(np.random.randint(7,10))

# commonStudents = []
# for i in range(C-1):
#     mini = min(numShortlists[i], numShortlists[i+1])
#     if 0 <= i and i < (C_1-1):
#         commonStudents.append(int(0.8*mini))
#     elif i < C_1:
#         commonStudents.append(int(0.50*mini))
#     elif i < (C_1 + C_2-1):
#         commonStudents.append(int(0.70*mini))
#     elif i < (C_1 + C_2 + 1):
#         commonStudents.append(int(mini))
#     else:
#         commonStudents.append(int(0.80*mini))
#     # p = np.random.rand()
#     # if p > 0:
#     #     commonStudents.append(np.random.randint(0, 6))
#     # else:
#     #     commonStudents.append(0)

# wb = load_workbook(F[6])
# ws = wb['Sheet1']
# curr_row = 2
# for i in range(2, 2+C):
#     ws.cell(row=i, column=1).value = firmName[i-2]
#     ws.cell(row=i, column=2).value = pocEmail[i-2]
# wb.save(F[6])

# print("\n", br)
# print("Number of shorlists:", numShortlists)
# print("Total shortlist: ", np.sum(np.array(numShortlists)))
# print("Overlapping students:", commonStudents)
# print(br)

# finalShortlist = []
# wb = load_workbook(filename = F[5])
# ws = wb['Sheet1']
# # curr_row = 2
# sum = 0
# for i in range(C):
#     sum += numShortlists[i]
#     if i < (C-1):
#         sum -= commonStudents[i]
# allShortlist = np.random.choice(regNum, sum, replace=False)
# curr = 0
# for i in range(2, 2+C):
#     tmp = allShortlist[curr:(curr+numShortlists[i-2])]
#     curr += numShortlists[i-2]
#     if i <= C:
#         curr -= commonStudents[i-2]
#     finalShortlist.append(list(set(tmp)))

# # for i in range(C):
# #     allShortlist = np.random.choice(regNum, numShortlists[i], replace=False)
# #     if i > 0 and i < 10:
# #         possible = int(min(numShortlists[i]/2, numShortlists[i-1]/2))
# #         allShortlist[0:possible] = finalShortlist[-1][0:possible]
# #     allShortlist = list(set(allShortlist))
# #     numShortlists[i] = len(allShortlist)
# #     finalShortlist.append(allShortlist)

# for i in range(2, 2+C):
#     for j in range(numShortlists[i-2]):
#         ws.cell(row=curr_row, column=1).value = firmName[i-2] + "(" + jobRole[i-2] + ")"
#         ws.cell(row=curr_row, column=2).value = finalShortlist[i-2][j]
#         ws.cell(row=curr_row, column=3).value = "SHORTLIST"
#         ws.cell(row=curr_row, column=4).value = 1
#         ws.cell(row=curr_row, column=5).value = "PI Round 1"
#         curr_row += 1
# wb.save(F[5])

# print("\nSimulation completed.")

# os.mkdir('/home/badola/APC/t' + str(counter))
# for i in range(len(F)):
#     copy2(F[i], '/home/badola/APC/t' + str(counter) + '/' + F[i])
# make_archive('/home/badola/APC/t' + str(counter), 'zip', root_dir='/home/badola/APC/t' + str(counter))

# print("\nDataset {} generation completed.\n\n".format(counter))
